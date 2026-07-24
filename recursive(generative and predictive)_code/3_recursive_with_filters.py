"""
Recursive MoLeR generation with broad HARD filters and graded SOFT filters.
1. Only clearly unusable structures are rejected by HARD filters.
2. Physicochemical properties are scored softly instead of requiring every
   preferred range to pass simultaneously.
3. UFF/MMFF/xTB failures do not reject a molecule.
4. Cheap 2D screening occurs before Mordred/xTB and CE prediction.
5. The MoLeR feedback loop is retained.
6. xTB temporary files are automatically deleted.
7. The overall top 20 unique molecules are saved to an Excel workbook.
8. The top 20 molecules and their properties are printed in the console.
The final Excel file is:
    top_20_generated_molecules.xlsx
"""
from __future__ import annotations
import math
import os
import re
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple
import joblib
import numpy as np
import pandas as pd
from joblib import Parallel, delayed
from mordred import Calculator, descriptors
from molecule_generation import load_model_from_directory
from openpyxl.styles import Alignment, Font, PatternFill
from rdkit import Chem, RDConfig
from rdkit.Chem import AllChem, Crippen, Descriptors, Lipinski, rdMolDescriptors
# ---------------------------------------------------------------------------
# Optional RDKit SA scorer
# ---------------------------------------------------------------------------
try:
    sys.path.append(os.path.join(RDConfig.RDContribDir, "SA_Score"))
    import sascorer
    SA_SCORER_AVAILABLE = True
except ImportError:
    print("WARNING: RDKit SA scorer is unavailable. "
        "SA score will be omitted from the soft score.")
    SA_SCORER_AVAILABLE = False
# ---------------------------------------------------------------------------
# Paths and recursive-generation settings
# ---------------------------------------------------------------------------
PREDICTIVE_MODEL_PATH = "saved_models/xgb_best_model.joblib"
PREPROCESSOR_PATH = "saved_models/preprocessor.joblib"
SOURCE_DATA_FILE = "cleaned_final_features.csv"
GENERATIVE_MODEL_DIR = "MODEL_DIR"
# Final Excel output
OUTPUT_EXCEL_PATH = "top_20_generated_molecules.xlsx"
FINAL_TOP_N = 20
# Start smaller while validating the workflow.
# Increase these values after confirming that the workflow runs correctly.
ITERATIONS = 10
BATCH_SIZE = 1000
TOP_K = 50
BASE_NOISE_SCALE = 0.50
N_JOBS = 5
RANDOM_SEED = 42
# Only this many molecules per iteration receive Mordred + xTB + CE prediction.
# Cheap 2D screening is still performed on the full BATCH_SIZE.
MAX_EXPENSIVE_EVALUATIONS = 400
# A molecule is preferred when its graded soft score is at least this value.
# This is not a hard rejection.
MIN_SOFT_SCORE = 45.0
# Keep some lower-soft-score structures for exploration rather than allowing
# the search to become trapped in one narrow chemical family.
EXPENSIVE_SELECTION_TOP_FRACTION = 0.75
# This does not reject a molecule. It is only used for reporting.
TARGET_CE_FOR_REPORTING = 90.0
XTB_CMD = shutil.which("xtb")
RNG = np.random.default_rng(RANDOM_SEED)
# ---------------------------------------------------------------------------
# Broad HARD limits
# Only reject clearly unusable structures.
# ---------------------------------------------------------------------------
HARD_FILTERS = {
    "allowed_elements": {"H","B","C","N","O","F","Si","P","S","Cl","Br","I",},
    "require_closed_shell": True,
    "formal_charge_min": -1,
    "formal_charge_max": 1,
    "mw_min": 30.0,
    "mw_max": 350.0,
    "heavy_atoms_min": 2,
    "heavy_atoms_max": 40,
    "maximum_fragments": 1,}
# ---------------------------------------------------------------------------
# SOFT ranges
# Each property receives a score from 0 to 1:
# 1.0 = inside the preferred range
# 0-1 = between the preferred and broad limits
# 0.0 = outside the broad range
# These are preference scores, not hard rejection rules.
# ---------------------------------------------------------------------------
SOFT_FILTERS = {
    "MolWt": {
        "preferred_min": 50.0,
        "preferred_max": 300.0,
        "broad_min": 30.0,
        "broad_max": 450.0,},
    "LogP": {
        "preferred_min": -2.0,
        "preferred_max": 1.0,
        "broad_min": -4.0,
        "broad_max": 2.5,},
    "TPSA": {
        "preferred_min": 20.0,
        "preferred_max": 100.0,
        "broad_min": 5.0,
        "broad_max": 150.0,},
    "HBD": {
        "preferred_min": 0.0,
        "preferred_max": 3.0,
        "broad_min": 0.0,
        "broad_max": 5.0,},
    "HBA": {
        "preferred_min": 1.0,
        "preferred_max": 6.0,
        "broad_min": 0.0,
        "broad_max": 10.0,},
    "NitrogenFraction": {
    "preferred_min": 0.0,
    "preferred_max": 0.20,
    "broad_min": 0.0,
    "broad_max": 0.30,},
    "RotatableBonds": {
        "preferred_min": 0.0,
        "preferred_max": 8.0,
        "broad_min": 0.0,
        "broad_max": 12.0,},
    "LargestRing": {
        "preferred_min": 0.0,
        "preferred_max": 6.0,
        "broad_min": 0.0,
        "broad_max": 8.0,},
    "SA_Score": {
        "preferred_min": 1.0,
        "preferred_max": 3.0,
        "broad_min": 1.0,
        "broad_max": 5.0,},
    "ESOL_logS": {
        "preferred_min": -1.5,
        "preferred_max": 0.5,
        "broad_min": -3.0,
        "broad_max": 1.5,},
    "FunctionalGroupMatchCount": {
        "preferred_min": 1.0,
        "preferred_max": 5.0,
        "broad_min": 0.0,
        "broad_max": 8.0,},
    "AromaticAtomFraction": {
    "preferred_min": 0.0,
    "preferred_max": 0.35,
    "broad_min": 0.0,
    "broad_max": 0.75,}}
RELEVANT_FUNCTIONAL_GROUPS = {
    "hydroxyl": "[OX2H]",
    "ether": "[OD2]([#6])[#6]",
    "carbonyl": "[CX3]=[OX1]",
    "amine": "[NX3;H2,H1,H0;+0]",
    "pyridine_like_N": "[nH0;+0]",
    "nitrile": "[CX2]#N",
    "sulfoxide_or_sulfone": "[SX3,SX4](=O)",
    "phosphoryl": "[PX4](=O)",
    "carboxylate_or_carboxylic_acid": "[CX3](=O)[OX1-,OX2H]",
    "amide": "[NX3][CX3](=[OX1])",}
# ---------------------------------------------------------------------------
# General utility functions
# ---------------------------------------------------------------------------
def finite_float(value: Any, default: float = 0.0) -> float:
    """Convert a value to a finite float or return the supplied default."""
    try:
        numeric = float(value)
        return numeric if math.isfinite(numeric) else default
    except Exception:
        return default
def aromatic_atom_fraction(mol):
    heavy_atoms = max(mol.GetNumHeavyAtoms(), 1)
    aromatic_atoms = sum(atom.GetIsAromatic() for atom in mol.GetAtoms())
    return aromatic_atoms / heavy_atoms
def canonicalize_smiles(smiles: str) -> Optional[str]:
    """Parse, sanitize, and canonicalize a SMILES string."""
    try:
        mol = Chem.MolFromSmiles(str(smiles))
        if mol is None:
            return None
        Chem.SanitizeMol(mol)
        return Chem.MolToSmiles(
            mol,
            canonical=True,
            isomericSmiles=True,)
    except Exception:
        return None
def largest_ring_size(mol: Chem.Mol) -> int:
    """Return the number of atoms in the molecule's largest ring."""
    try:
        rings = mol.GetRingInfo().AtomRings()
        return max((len(ring) for ring in rings), default=0)
    except Exception:
        return 0
def nitrogen_fraction(mol: Chem.Mol) -> float:
    heavy_atoms = max(mol.GetNumHeavyAtoms(), 1)
    nitrogen_atoms = sum(
        atom.GetSymbol() == "N"
        for atom in mol.GetAtoms())
    return nitrogen_atoms / heavy_atoms
def is_closed_shell(mol: Chem.Mol) -> bool:
    """Return True when no atom has radical electrons."""
    return all(
        atom.GetNumRadicalElectrons() == 0
        for atom in mol.GetAtoms())
def allowed_elements_pass(
    mol: Chem.Mol,
) -> Tuple[bool, List[str]]:
    """Check whether every element is included in the allowed element set."""
    present = {
        atom.GetSymbol()
        for atom in mol.GetAtoms()}
    disallowed = sorted(
        present - HARD_FILTERS["allowed_elements"])
    return len(disallowed) == 0, disallowed
def smarts_match_summary(
    mol: Chem.Mol,
    smarts_map: Dict[str, str],
) -> Tuple[List[str], int]:
    """Return functional-group names and the total number of matches."""
    matched_names: List[str] = []
    total_matches = 0
    for name, smarts in smarts_map.items():
        pattern = Chem.MolFromSmarts(smarts)
        if pattern is None:
            continue
        matches = mol.GetSubstructMatches(pattern)
        if matches:
            matched_names.append(name)
            total_matches += len(matches)
    return matched_names, total_matches
def estimate_esol_logS(mol: Chem.Mol) -> float:
    """Calculate a Delaney-style ESOL log10(mol/L) estimate."""
    mw = Descriptors.MolWt(mol)
    logp = Crippen.MolLogP(mol)
    rotors = Lipinski.NumRotatableBonds(mol)
    heavy_atoms = max(mol.GetNumHeavyAtoms(), 1)
    aromatic_atoms = sum(
        atom.GetIsAromatic()
        for atom in mol.GetAtoms())
    aromatic_proportion = aromatic_atoms / heavy_atoms
    return (
        0.16
        - 0.63 * logp
        - 0.0062 * mw
        + 0.066 * rotors
        - 0.74 * aromatic_proportion)
def compute_rdkit_properties(
    mol: Chem.Mol,
) -> Dict[str, float]:
    """Calculate inexpensive RDKit properties."""
    return {
        "MolWt": float(Descriptors.MolWt(mol)),
        "LogP": float(Crippen.MolLogP(mol)),
        "TPSA": float(rdMolDescriptors.CalcTPSA(mol)),
        "HBD": float(Lipinski.NumHDonors(mol)),
        "HBA": float(Lipinski.NumHAcceptors(mol)),
        "RotatableBonds": float(
            Lipinski.NumRotatableBonds(mol)),
        "HeavyAtoms": float(mol.GetNumHeavyAtoms()),
        "NitrogenFraction": float(nitrogen_fraction(mol)),
        "FormalCharge": float(Chem.GetFormalCharge(mol)),
        "LargestRing": float(largest_ring_size(mol)),
        "ESOL_logS": float(estimate_esol_logS(mol)),
        "FragmentCount": float(
            len(Chem.GetMolFrags(mol))),}
def graded_range_score(
    value: Any,
    *,
    preferred_min: float,
    preferred_max: float,
    broad_min: float,
    broad_max: float,
) -> Optional[float]:
    """
    Return a continuous preference score between 0 and 1.
    Returns None when the value is missing or invalid.
    """
    try:
        x = float(value)
    except Exception:
        return None
    if not math.isfinite(x):
        return None
    if preferred_min <= x <= preferred_max:
        return 1.0
    if x < preferred_min:
        if x <= broad_min:
            return 0.0
        denominator = preferred_min - broad_min
        if denominator <= 0:
            return 0.0
        return float(
            (x - broad_min) / denominator)
    if x > preferred_max:
        if x >= broad_max:
            return 0.0
        denominator = broad_max - preferred_max
        if denominator <= 0:
            return 0.0
        return float(
            (broad_max - x) / denominator)
    return 0.0
def calculate_soft_score(
    row: Dict[str, Any],
) -> Tuple[float, Dict[str, float], List[str]]:
    """Calculate the overall 0-100 graded chemistry score."""
    component_scores: Dict[str, float] = {}
    failed_preferred: List[str] = []
    for name, limits in SOFT_FILTERS.items():
        score = graded_range_score(
            row.get(name),
            **limits,)
        if score is None:
            continue
        component_scores[name] = score
        if score < 1.0:
            failed_preferred.append(name)
    if not component_scores:
        return 0.0, component_scores, failed_preferred
    soft_score = 100.0 * float(
        np.mean(list(component_scores.values())))
    return soft_score, component_scores, failed_preferred
# ---------------------------------------------------------------------------
# Cheap screening
# No Mordred, no 3D geometry, and no xTB.
# ---------------------------------------------------------------------------
def cheap_screen_single_molecule(
    smiles: str,
    index: int,
    iteration_number: int,
) -> Dict[str, Any]:
    """Run inexpensive hard filtering and soft scoring."""
    row: Dict[str, Any] = {
        "Input_SMILES": smiles,
        "CandidateIndex": index,
        "Iteration": iteration_number,}
    hard_failures: List[str] = []
    canonical = canonicalize_smiles(smiles)
    if canonical is None:
        row.update({
            "SMILES": smiles,
            "HardPass": False,
            "HardFailureReasons": (
                "invalid_or_unsanitizable_structure"
            ),
            "SoftScore": 0.0,})
        return row
    row["SMILES"] = canonical
    mol = Chem.MolFromSmiles(canonical)
    if mol is None:
        row.update({
            "HardPass": False,
            "HardFailureReasons": (
                "invalid_structure_after_canonicalization"),
            "SoftScore": 0.0,})
        return row
    properties = compute_rdkit_properties(mol)
    row.update(properties)
    # Closed-shell check
    if (
        HARD_FILTERS["require_closed_shell"]
        and not is_closed_shell(mol)):
        hard_failures.append(
            "contains_radical_electrons")
    # Allowed-element check
    elements_ok, disallowed = allowed_elements_pass(mol)
    row["DisallowedElements"] = ",".join(disallowed)
    if not elements_ok:
        hard_failures.append(
            f"disallowed_elements:{','.join(disallowed)}")
    # Formal-charge check
    formal_charge = int(properties["FormalCharge"])
    if not (
        HARD_FILTERS["formal_charge_min"]
        <= formal_charge
        <= HARD_FILTERS["formal_charge_max"]):
        hard_failures.append(
            "formal_charge_outside_-2_to_+2")
    # Molecular-weight check
    if not (
        HARD_FILTERS["mw_min"]
        <= properties["MolWt"]
        <= HARD_FILTERS["mw_max"]):
        hard_failures.append(
            "molecular_weight_outside_30_to_450")
    # Heavy-atom check
    if not (
        HARD_FILTERS["heavy_atoms_min"]
        <= properties["HeavyAtoms"]
        <= HARD_FILTERS["heavy_atoms_max"]):
        hard_failures.append(
            "heavy_atom_count_outside_2_to_40")
    # Disconnected-fragment check
    if (
        properties["FragmentCount"]
        > HARD_FILTERS["maximum_fragments"]):
        hard_failures.append(
            "multiple_disconnected_fragments")
    # Functional-group matching
    functional_groups, group_count = smarts_match_summary(
        mol,
        RELEVANT_FUNCTIONAL_GROUPS,)
    row["MatchedFunctionalGroups"] = ",".join(
        functional_groups)
    row["FunctionalGroupMatchCount"] = float(
        group_count)
    # Synthetic accessibility
    sa_score = float("nan")
    if SA_SCORER_AVAILABLE:
        try:
            sa_score = float(
                sascorer.calculateScore(mol))
        except Exception:
            pass
    row["SA_Score"] = sa_score
    # Graded soft scoring
    soft_score, components, failed_preferred = (
        calculate_soft_score(row))
    row["SoftScore"] = soft_score
    row["SoftPreferredChecksPassed"] = int(
        sum(
            score >= 0.999
            for score in components.values()))
    row["SoftChecksEvaluated"] = int(
        len(components))
    row["SoftFailedPreferredRanges"] = ",".join(
        failed_preferred)
    for name, score in components.items():
        row[f"SoftComponent_{name}"] = score
    row["HardPass"] = len(hard_failures) == 0
    row["HardFailureReasons"] = ";".join(
        hard_failures)
    return row
# ---------------------------------------------------------------------------
# 3D geometry and xTB helpers
# Failure does not reject a molecule.
# ---------------------------------------------------------------------------
def generate_3d_xyz(
    smiles: str,
    xyz_path: Path,
) -> Tuple[bool, str]:
    """Generate an initial 3D structure and write it as an XYZ file."""
    mol = Chem.MolFromSmiles(smiles)
    if mol is None:
        return False, "invalid"
    mol = Chem.AddHs(mol)
    params = AllChem.ETKDGv3()
    params.randomSeed = RANDOM_SEED
    try:
        status = AllChem.EmbedMolecule(
            mol,
            params,)
    except Exception:
        return False, "embedding_exception"
    if status != 0:
        return False, "embedding_failed"
    optimization_method = "embedded_unoptimized"
    try:
        if AllChem.MMFFHasAllMoleculeParams(mol):
            AllChem.MMFFOptimizeMolecule(
                mol,
                maxIters=300,)
            optimization_method = "MMFF"
        elif AllChem.UFFHasAllMoleculeParams(mol):
            AllChem.UFFOptimizeMolecule(
                mol,
                maxIters=300,)
            optimization_method = "UFF"
    except Exception:
        optimization_method = "embedded_unoptimized"
    try:
        conf = mol.GetConformer()
        with xyz_path.open(
            "w",
            encoding="utf-8",
        ) as handle:
            handle.write(
                f"{mol.GetNumAtoms()}\n")
            handle.write(
                f"{smiles}\n")
            for atom in mol.GetAtoms():
                point = conf.GetAtomPosition(
                    atom.GetIdx())
                handle.write(
                    f"{atom.GetSymbol()} "
                    f"{point.x:.8f} "
                    f"{point.y:.8f} "
                    f"{point.z:.8f}\n")
    except Exception:
        return False, "xyz_write_failed"
    return True, optimization_method
def run_xtb(
    work_dir: Path,
    xyz_path: Path,
    log_path: Path,
) -> bool:
    """Run GFN2-xTB inside a temporary directory."""
    if XTB_CMD is None:
        return False
    command = [
        XTB_CMD,
        xyz_path.name,
        "--gfn",
        "2",
        "--parallel",
        "1",]
    try:
        with log_path.open(
            "w",
            encoding="utf-8",
        ) as log_handle:
            subprocess.run(
                command,
                cwd=work_dir,
                stdout=log_handle,
                stderr=subprocess.STDOUT,
                check=True,
                timeout=90,)
        return True
    except (
        subprocess.CalledProcessError,
        subprocess.TimeoutExpired,
        OSError,):
        return False
def first_regex_float(
    text: str,
    patterns: Sequence[str],
) -> float:
    """Return the first successfully parsed floating-point regex result."""
    for pattern in patterns:
        match = re.search(
            pattern,
            text,
            flags=re.IGNORECASE | re.MULTILINE,)
        if match:
            try:
                return float(match.group(1))
            except Exception:
                continue
    return float("nan")
def parse_xtb_log(
    log_path: Path,
) -> Dict[str, float]:
    """Extract HOMO, LUMO, energy gap, and dipole from an xTB log."""
    if not log_path.exists():
        return {}
    text = log_path.read_text(
        errors="ignore")
    homo = first_regex_float(
        text,
        [r"\(HOMO\)\s+[-\d.]+\s+([-\d.]+)",
        r"HOMO\s+energy\s*[:=]\s*([-\d.]+)\s*eV",],)
    lumo = first_regex_float(
        text,
        [r"\(LUMO\)\s+[-\d.]+\s+([-\d.]+)",
         r"LUMO\s+energy\s*[:=]\s*([-\d.]+)\s*eV",],)
    gap = first_regex_float(
        text,
        [r"HOMO-LUMO\s+GAP\s+([-\d.]+)\s+eV",],)
    dipole = first_regex_float(
        text,
        [(r"molecular dipole:[\s\S]*?"
        r"full:\s+[-\d.]+\s+[-\d.]+\s+[-\d.]+"
        r"\s+([-\d.]+)"),],)
    return {
        "HOMO_eV": homo,
        "LUMO_eV": lumo,
        "HOMO_LUMO_Gap_eV": gap,
        "Dipole_D": dipole,}
def calculate_mordred(
    mol: Chem.Mol,
    calculator: Calculator,
) -> Dict[str, float]:
    """Calculate finite Mordred descriptor values."""
    result = calculator(mol)
    output: Dict[str, float] = {}
    for descriptor_name, value in result.asdict().items():
        try:
            numeric = float(value)
            if math.isfinite(numeric):
                output[str(descriptor_name)] = numeric
        except Exception:
            continue
    return output
def build_aligned_query(
    feature_names: Sequence[str],
    combined_features: Dict[str, Any],
    mordred_map: Dict[str, str],
) -> pd.DataFrame:
    """Align generated molecular features with the trained model columns."""
    query = pd.DataFrame(
        0.0,
        index=[0],
        columns=list(feature_names),)
    for column in feature_names:
        candidates = [
            column,
            mordred_map.get(column, column),]
        for candidate in candidates:
            if candidate in combined_features:
                query.at[0, column] = finite_float(
                    combined_features[candidate])
                break
    return query
# ---------------------------------------------------------------------------
# Expensive evaluation and CE prediction
# ---------------------------------------------------------------------------
def evaluate_expensive_and_predict(
    cheap_row: Dict[str, Any],
    *,
    xgb_model: Any,
    preprocessor: Any,
    feature_names: Sequence[str],
    mordred_calc: Calculator,
    mordred_map: Dict[str, str],
) -> Dict[str, Any]:
    """Calculate Mordred/xTB features and predict CE."""
    row = dict(cheap_row)
    canonical = row["SMILES"]
    mol = Chem.MolFromSmiles(canonical)
    if mol is None:
        row.update({
            "Predicted_CE": float("nan"),
            "CEPredictionAvailable": False,
            "XTBSuccess": False,})
        return row
    # Mordred calculation
    try:
        mordred_features = calculate_mordred(
            mol,
            mordred_calc,)
    except Exception:
        mordred_features = {}
    # Default xTB properties
    xtb_properties = {
        "HOMO_eV": float("nan"),
        "LUMO_eV": float("nan"),
        "HOMO_LUMO_Gap_eV": float("nan"),
        "Dipole_D": float("nan"),}
    xtb_success = False
    geometry_method = "not_attempted"
    # xTB side files are created here and automatically deleted.
    with tempfile.TemporaryDirectory(
        prefix="moler_xtb_"
    ) as temporary_directory:
        work_dir = Path(temporary_directory)
        xyz_path = work_dir / "candidate.xyz"
        log_path = work_dir / "xtb.log"
        geometry_success, geometry_method = (
            generate_3d_xyz(
                canonical,
                xyz_path,))
        if geometry_success:
            xtb_success = run_xtb(
                work_dir,
                xyz_path,
                log_path,)
            if xtb_success:
                xtb_properties.update(
                    parse_xtb_log(log_path))
    row["GeometryMethod"] = geometry_method
    row["XTBSuccess"] = bool(xtb_success)
    row.update(xtb_properties)
    combined_features = {
        **row,
        **mordred_features,
        **xtb_properties,}
    aligned_query = build_aligned_query(
        feature_names,
        combined_features,
        mordred_map,)
    try:
        processed = preprocessor.transform(
            aligned_query)
        predicted_ce = float(
            np.asarray(
                xgb_model.predict(processed)
            ).reshape(-1)[0])
    except Exception as exc:
        predicted_ce = float("nan")
        row["CEPredictionError"] = str(exc)
    row["Predicted_CE"] = predicted_ce
    row["CEPredictionAvailable"] = math.isfinite(
        predicted_ce)
    row["MeetsReportedCETarget"] = (
        math.isfinite(predicted_ce)
        and predicted_ce >= TARGET_CE_FOR_REPORTING)
    # CE remains the main recursive selection target.
    # Soft chemistry score is used as a tie-breaker.
    row["SelectionScore"] = (
        predicted_ce
        if math.isfinite(predicted_ce)
        else -math.inf)
    return row
# ---------------------------------------------------------------------------
# Sampling and candidate-selection helpers
# ---------------------------------------------------------------------------
def sample_unique(
    model: Any,
    number: int,
    exclude: Optional[set] = None,
) -> List[str]:
    """Sample unique canonical SMILES from MoLeR."""
    exclude = exclude or set()
    collected: Dict[str, None] = {}
    attempts = 0
    max_attempts = 100
    while (
        len(collected) < number
        and attempts < max_attempts):
        attempts += 1
        requested = max(
            5,
            number - len(collected),)
        try:
            sampled_smiles = model.sample(
                requested)
        except Exception as exc:
            print(
                f"MoLeR sampling error: {exc}",
                flush=True,)
            break
        for smiles in sampled_smiles:
            canonical = canonicalize_smiles(smiles)
            if (
                canonical
                and canonical not in exclude):
                collected[canonical] = None
    return list(collected)[:number]
def select_for_expensive_evaluation(
    hard_pass_df: pd.DataFrame,
    maximum_count: int,
) -> pd.DataFrame:
    """
    Retain mostly high-soft-score molecules together with a random
    exploration subset.
    This prevents the soft ranges from becoming hidden hard filters.
    """
    if hard_pass_df.empty:
        return hard_pass_df
    ordered = hard_pass_df.sort_values(
        "SoftScore",
        ascending=False,
    ).reset_index(drop=True)
    if len(ordered) <= maximum_count:
        return ordered
    exploitation_count = int(
        maximum_count
        * EXPENSIVE_SELECTION_TOP_FRACTION)
    exploration_count = (
        maximum_count - exploitation_count)
    exploitation = ordered.head(
        exploitation_count)
    remaining = ordered.iloc[
        exploitation_count:]
    if (
        exploration_count > 0
        and not remaining.empty):
        chosen_positions = RNG.choice(len(remaining), size=min(exploration_count, len(remaining),), replace=False,)
        exploration = remaining.iloc[chosen_positions]
        selected = pd.concat([exploitation, exploration], ignore_index=True,)
    else:
        selected = exploitation.copy()
    return (
        selected
        .drop_duplicates(subset=["SMILES"])
        .reset_index(drop=True))
def print_filter_summary(
    screen_df: pd.DataFrame,
) -> None:
    """Print hard-filter and soft-score summary statistics."""
    total = len(screen_df)
    if "HardPass" in screen_df.columns:
        hard_pass_count = int(
            screen_df["HardPass"]
            .fillna(False)
            .astype(bool)
            .sum())
    else:
        hard_pass_count = 0
    if "SoftScore" in screen_df.columns:
        soft_values = pd.to_numeric(
            screen_df["SoftScore"],
            errors="coerce",)
        soft_preferred_count = int((soft_values >= MIN_SOFT_SCORE).sum())
    else:
        soft_preferred_count = 0
    print(f"Cheap-screened molecules: {total}", flush=True,)
    print(f"Passed broad HARD filters: "f"{hard_pass_count}/{total}", flush=True,)
    print(f"Reached preferred SOFT score >= "f"{MIN_SOFT_SCORE:.1f}: "f"{soft_preferred_count}/{total}", flush=True,)
    if ("HardFailureReasons"
        in screen_df.columns):
        hard_reasons = (
            screen_df.loc[
                screen_df["HardPass"] == False,
                "HardFailureReasons",]
            .fillna("")
            .str.split(";")
            .explode())
        hard_reasons = (
            hard_reasons[
                hard_reasons != ""]
            .value_counts())
        if not hard_reasons.empty:
            print(
                "Most common HARD failures:",
                flush=True,)
            print(
                hard_reasons
                .head(8)
                .to_string(),
                flush=True,)
    if ("SoftFailedPreferredRanges"
        in screen_df.columns):
        soft_reasons = (
            screen_df[
                "SoftFailedPreferredRanges"]
            .fillna("")
            .str.split(",")
            .explode())
        soft_reasons = (
            soft_reasons[
                soft_reasons != ""]
            .value_counts())
        if not soft_reasons.empty:
            print(
                "Most common properties outside "
                "the preferred SOFT range:",
                flush=True,)
            print(
                soft_reasons
                .head(8)
                .to_string(),
                flush=True,)
# ---------------------------------------------------------------------------
# Safe DataFrame-column helpers for Excel preparation
# ---------------------------------------------------------------------------
def numeric_column(
    dataframe: pd.DataFrame,
    column_name: str,
    default: float = float("nan"),
) -> pd.Series:
    """Return a numeric Series even when a column is missing."""
    if column_name in dataframe.columns:
        return pd.to_numeric(
            dataframe[column_name],
            errors="coerce",)
    return pd.Series(
        [default] * len(dataframe),
        index=dataframe.index,
        dtype=float,)
def text_column(
    dataframe: pd.DataFrame,
    column_name: str,
    default: str = "",
) -> pd.Series:
    """Return a text Series even when a column is missing."""
    if column_name in dataframe.columns:
        return (
            dataframe[column_name]
            .fillna(default)
            .astype(str))
    return pd.Series(
        [default] * len(dataframe),
        index=dataframe.index,
        dtype=str,)
def boolean_column(
    dataframe: pd.DataFrame,
    column_name: str,
    default: bool = False,
) -> pd.Series:
    """Return a Boolean Series even when a column is missing."""
    if column_name in dataframe.columns:
        return (
            dataframe[column_name]
            .fillna(default)
            .astype(bool))
    return pd.Series(
        [default] * len(dataframe),
        index=dataframe.index,
        dtype=bool,)
# ---------------------------------------------------------------------------
# Prepare and save the final top 20 Excel table
# ---------------------------------------------------------------------------
def prepare_top_molecules_table(
    all_scored_molecules: List[Dict[str, Any]],
    final_top_n: int,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    Combine results from every iteration, remove duplicate SMILES,
    and select the overall best unique molecules.
    """
    all_results_df = pd.DataFrame(
        all_scored_molecules)
    if all_results_df.empty:
        return pd.DataFrame(), pd.DataFrame()
    required_columns = [
        "SMILES",
        "Predicted_CE",
        "SoftScore",]
    missing_required = [
        column
        for column in required_columns
        if column not in all_results_df.columns]
    if missing_required:
        raise KeyError(
            "The generated results are missing required "
            f"columns: {missing_required}")
    all_results_df["Predicted_CE"] = pd.to_numeric(
        all_results_df["Predicted_CE"],
        errors="coerce",)
    all_results_df["SoftScore"] = pd.to_numeric(
        all_results_df["SoftScore"],
        errors="coerce",)
    all_results_df = all_results_df.dropna(
        subset=[
            "SMILES",
            "Predicted_CE",
            "SoftScore",]
    ).copy()
    if all_results_df.empty:
        return pd.DataFrame(), pd.DataFrame()
    # Rank first by predicted CE and then by soft chemistry score.
    all_results_df = all_results_df.sort_values(
        [
            "Predicted_CE",
            "SoftScore",
        ],
        ascending=[
            False,
            False,
        ],)
    # The same molecule might appear in several iterations.
    # Keep its highest-ranked occurrence.
    all_results_df = (
        all_results_df
        .drop_duplicates(
            subset=["SMILES"],
            keep="first",)
        .reset_index(drop=True))
    # Prefer molecules meeting the minimum soft score.
    preferred_results_df = all_results_df[
        all_results_df["SoftScore"]
        >= MIN_SOFT_SCORE
    ].copy()
    # If fewer than 20 meet the preferred score, fill the remaining
    # positions with the highest-ranked hard-pass molecules.
    if len(preferred_results_df) >= final_top_n:
        top_df = preferred_results_df.head(
            final_top_n
        ).copy()
    else:
        preferred_smiles = set(
            preferred_results_df["SMILES"])
        remaining_df = all_results_df[
            ~all_results_df["SMILES"].isin(
                preferred_smiles)
        ].copy()
        number_needed = (
            final_top_n
            - len(preferred_results_df))
        top_df = pd.concat(
            [
                preferred_results_df,
                remaining_df.head(number_needed),],
            ignore_index=True,)
    # Re-sort the combined final set.
    top_df = top_df.sort_values(
        [
            "Predicted_CE",
            "SoftScore",],
        ascending=[
            False,
            False,],
    ).reset_index(drop=True)
    top_df.insert(
        0,
        "Rank",
        np.arange(
            1,
            len(top_df) + 1,),)
    # Readable output table for Excel.
    excel_df = pd.DataFrame({
        "Rank": top_df["Rank"],
        "SMILES": text_column(top_df,"SMILES",),
        "Predicted CE (%)": numeric_column(top_df,"Predicted_CE",).round(2),
        "Soft chemistry score (/100)": numeric_column(top_df,"SoftScore",).round(2),
        "Molecular weight": numeric_column(top_df,"MolWt",).round(2),
        "LogP": numeric_column(top_df,"LogP",).round(2),
        "TPSA": numeric_column(top_df,"TPSA",).round(2),
        "ESOL logS": numeric_column(top_df,"ESOL_logS",).round(2),
        "SA score": numeric_column(top_df,"SA_Score",).round(2),
        "xTB success": boolean_column(top_df,"XTBSuccess",),
        "HOMO (eV)": numeric_column(top_df,"HOMO_eV",).round(4),
        "LUMO (eV)": numeric_column(top_df,"LUMO_eV",).round(4),
        "HOMO-LUMO gap (eV)": numeric_column(top_df,"HOMO_LUMO_Gap_eV",).round(4),
        "Dipole (D)": numeric_column(top_df,"Dipole_D",).round(4),
        "Formal charge": numeric_column(top_df,"FormalCharge",),
        "Heavy atoms": numeric_column(top_df,"HeavyAtoms",),
        "Largest ring": numeric_column(top_df,"LargestRing",),
        "Rotatable bonds": numeric_column(top_df,"RotatableBonds",),
        "H-bond donors": numeric_column(top_df,"HBD",),
        "H-bond acceptors": numeric_column(top_df,"HBA",),
        "Matched functional groups": text_column(top_df,"MatchedFunctionalGroups",),
        "Functional-group match count": numeric_column(top_df,"FunctionalGroupMatchCount",),
        "Preferred soft checks passed": numeric_column(top_df,"SoftPreferredChecksPassed",),
        "Soft checks evaluated": numeric_column(top_df,"SoftChecksEvaluated",),
        "Properties outside preferred range": text_column(top_df,"SoftFailedPreferredRanges",),
        "Geometry method": text_column(top_df,"GeometryMethod",),
        "Iteration found": numeric_column(top_df,"Iteration",),
        "Candidate index": numeric_column(top_df,"CandidateIndex",),
        "Reached CE reporting target": boolean_column(top_df,"MeetsReportedCETarget",),
    })
    return top_df, excel_df
def save_top_molecules_excel(
    excel_df: pd.DataFrame,
    output_path: str,
) -> None:
    """Save and format the final top-molecule Excel workbook."""
    output_file = Path(output_path)
    output_file.parent.mkdir(
        parents=True,
        exist_ok=True,)
    with pd.ExcelWriter(
        output_file,
        engine="openpyxl",
    ) as writer:
        excel_df.to_excel(
            writer,
            sheet_name="Top 20 Molecules",
            index=False,)
        worksheet = writer.sheets[
            "Top 20 Molecules"]
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = (
            worksheet.dimensions)
        # Header formatting
        header_fill = PatternFill(
            fill_type="solid",
            fgColor="1F4E78",)
        header_font = Font(
            color="FFFFFF",
            bold=True,)
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True, )
        # General cell alignment
        for row in worksheet.iter_rows(
            min_row=2):
            for cell in row:
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )
        # Automatically adjust column widths.
        for column_cells in worksheet.columns:
            maximum_length = 0
            column_letter = (
                column_cells[0].column_letter)
            for cell in column_cells:
                value = (""if cell.value is None
                    else str(cell.value))
                maximum_length = max(
                    maximum_length,
                    len(value),)
            worksheet.column_dimensions[column_letter].width = min(maximum_length + 2, 45,)
        # Make the SMILES column wider.
        worksheet.column_dimensions["B"].width = 45
        # Format rank and iteration-like columns as integers.
        integer_headers = {
            "Rank",
            "Heavy atoms",
            "Largest ring",
            "Rotatable bonds",
            "H-bond donors",
            "H-bond acceptors",
            "Functional-group match count",
            "Preferred soft checks passed",
            "Soft checks evaluated",
            "Iteration found",
            "Candidate index",}
        header_positions = {cell.value: cell.column for cell in worksheet[1]}
        for header in integer_headers:
            column_index = header_positions.get(
                header)
            if column_index is None:
                continue
            for row_index in range(
                2,
                worksheet.max_row + 1,):
                worksheet.cell(
                    row=row_index,
                    column=column_index,
                ).number_format = "0"
        # Format major numeric values.
        two_decimal_headers = {
            "Predicted CE (%)",
            "Soft chemistry score (/100)",
            "Molecular weight",
            "LogP",
            "TPSA",
            "ESOL logS",
            "SA score",}
        for header in two_decimal_headers:
            column_index = header_positions.get(
                header)
            if column_index is None:
                continue
            for row_index in range(2,
                worksheet.max_row + 1,):
                worksheet.cell(
                    row=row_index,
                    column=column_index,
                ).number_format = "0.00"
def print_top_molecules(
    excel_df: pd.DataFrame,
) -> None:
    """Print the requested properties for each final molecule."""
    print(
        f"\nTOP {len(excel_df)} GENERATED MOLECULES",
        flush=True,)
    for _, molecule in excel_df.iterrows():
        print(
            "\n" + "-" * 70,
            flush=True,)
        print(
            f"Rank: {int(molecule['Rank'])}",
            flush=True,)
        print(
            f"SMILES: {molecule['SMILES']}",
            flush=True,)
        print(
            f"Predicted CE: "
            f"{finite_float(molecule['Predicted CE (%)'], float('nan')):.2f}%",
            flush=True,)
        print(
            "Soft chemistry score: "
            f"{finite_float(molecule['Soft chemistry score (/100)'], float('nan')):.1f}/100",
            flush=True,)
        print(
            "Molecular weight: "
            f"{finite_float(molecule['Molecular weight'], float('nan')):.2f}",
            flush=True,)
        print(
            f"LogP: "
            f"{finite_float(molecule['LogP'], float('nan')):.2f}",
            flush=True,)
        print(
            f"TPSA: "
            f"{finite_float(molecule['TPSA'], float('nan')):.2f}",
            flush=True,)
        print(
            f"ESOL logS: "
            f"{finite_float(molecule['ESOL logS'], float('nan')):.2f}",
            flush=True,)
        print(
            f"SA score: "
            f"{finite_float(molecule['SA score'], float('nan')):.2f}",
            flush=True,)
        print(
            f"xTB success: "
            f"{bool(molecule['xTB success'])}",
            flush=True,)
# ---------------------------------------------------------------------------
# Main recursive MoLeR loop
# ---------------------------------------------------------------------------
def main() -> Optional[pd.DataFrame]:
    """Run recursive generation and save the final top 20 molecules."""
    print(
        "--- Loading predictive model and preprocessing pipeline ---",
        flush=True,
    )
    # Validate paths before beginning the expensive workflow.
    required_paths = {
        "Predictive model": Path(
            PREDICTIVE_MODEL_PATH),
        "Preprocessor": Path(
            PREPROCESSOR_PATH),
        "Source feature file": Path(
            SOURCE_DATA_FILE),
        "MoLeR model directory": Path(
            GENERATIVE_MODEL_DIR),}
    missing_paths = [
        f"{name}: {path}"
        for name, path in required_paths.items()
        if not path.exists()]
    if missing_paths:
        raise FileNotFoundError(
            "The following required paths do not exist:\n"
            + "\n".join(missing_paths))
    if XTB_CMD is None:
        print(
            "WARNING: xTB was not found in PATH. "
            "The script will continue, but xTB success will be False "
            "and xTB properties may remain blank.",
            flush=True,)
    else:
        print(f"xTB executable found: {XTB_CMD}",
            flush=True,)
    xgb_model = joblib.load(
        PREDICTIVE_MODEL_PATH)
    preprocessor = joblib.load(
        PREPROCESSOR_PATH)
    source_df = pd.read_csv(
        SOURCE_DATA_FILE)
    target = "CE_aver. (%)"
    drop_columns = [
        "CE_1 (%)",
        "CE_2 (%)",
        "CE_3 (%)",
        target,
        "Additive_SMILES",
        "id",]
    existing_drop_columns = [
        column
        for column in drop_columns
        if column in source_df.columns]
    feature_names = list(
        source_df.drop(
            columns=existing_drop_columns
        ).columns)
    if not feature_names:
        raise ValueError(
            "No model input feature columns were found "
            "in the source feature file.")
    print(
        f"Predictive feature count: {len(feature_names)}",
        flush=True,)
    mordred_calc = Calculator(
        descriptors,
        ignore_3D=True,)
    # Map common RDKit column names to possible Mordred names.
    mordred_map = {
        "MolWt": "MW",
        "LogP": "LogP",
        "TPSA": "TopoPSA",
        "HBD": "nHBDon",
        "HBA": "nHBAcc",}
    # Store every valid scored candidate from every iteration.
    all_scored_molecules: List[
        Dict[str, Any]] = []
    previous_best_ce: Optional[float] = None
    current_noise = BASE_NOISE_SCALE
    print(
        "\n--- Starting recursive MoLeR soft/hard-filter pipeline ---",
        flush=True,)
    print(
        "xTB temporary files will be deleted automatically.\n"
        f"The final top {FINAL_TOP_N} unique molecules will be saved to:\n"
        f"{Path(OUTPUT_EXCEL_PATH).resolve()}\n",
        flush=True,
    )
    with load_model_from_directory(
        GENERATIVE_MODEL_DIR
    ) as model:
        print(
            "MoLeR loaded successfully.",
            flush=True,
        )
        current_smiles = sample_unique(
            model,
            BATCH_SIZE,
        )
        if not current_smiles:
            raise RuntimeError(
                "MoLeR did not generate any valid initial molecules."
            )
        previously_seen = set(
            current_smiles
        )
        for iteration in range(
            1,
            ITERATIONS + 1,
        ):
            print(
                "\n" + "=" * 78,
                flush=True,
            )
            print(
                f"ITERATION {iteration}/{ITERATIONS} | "
                f"noise={current_noise:.2f} | "
                f"generated={len(current_smiles)}",
                flush=True,
            )
            print(
                "Stage 1/3: broad HARD filtering "
                "and graded SOFT scoring...",
                flush=True,
            )
            cheap_rows = Parallel(
                n_jobs=N_JOBS,
                verbose=5,
            )(
                delayed(
                    cheap_screen_single_molecule
                )(
                    smiles,
                    index,
                    iteration,
                )
                for index, smiles in enumerate(
                    current_smiles
                )
            )
            screen_df = pd.DataFrame(
                cheap_rows
            )
            if screen_df.empty:
                print(
                    "No molecules could be screened; "
                    "sampling a fresh batch.",
                    flush=True,
                )
                current_smiles = sample_unique(
                    model,
                    BATCH_SIZE,
                    exclude=previously_seen,
                )
                previously_seen.update(
                    current_smiles
                )
                continue
            if "SMILES" not in screen_df.columns:
                print(
                    "No valid SMILES column was produced; "
                    "sampling a fresh batch.",
                    flush=True,
                )
                current_smiles = sample_unique(
                    model,
                    BATCH_SIZE,
                    exclude=previously_seen,
                )
                previously_seen.update(
                    current_smiles
                )
                continue
            screen_df = (
                screen_df
                .drop_duplicates(
                    subset=["SMILES"],
                    keep="first",
                )
                .reset_index(drop=True)
            )
            print_filter_summary(
                screen_df
            )
            hard_pass_df = screen_df[
                screen_df["HardPass"] == True
            ].copy()
            if hard_pass_df.empty:
                print(
                    "No molecule passed the broad HARD filters; "
                    "sampling a fresh batch.",
                    flush=True,
                )
                current_smiles = sample_unique(
                    model,
                    BATCH_SIZE,
                    exclude=previously_seen,
                )
                previously_seen.update(
                    current_smiles
                )
                continue
            expensive_input_df = (
                select_for_expensive_evaluation(
                    hard_pass_df,
                    maximum_count=(
                        MAX_EXPENSIVE_EVALUATIONS
                    ),
                )
            )
            print(
                "Stage 2/3: Mordred + optional xTB + "
                f"CE prediction for {len(expensive_input_df)} molecules...",
                flush=True,
            )
            scored_rows = Parallel(
                n_jobs=N_JOBS,
                verbose=10,
            )(
                delayed(
                    evaluate_expensive_and_predict
                )(
                    row,
                    xgb_model=xgb_model,
                    preprocessor=preprocessor,
                    feature_names=feature_names,
                    mordred_calc=mordred_calc,
                    mordred_map=mordred_map,
                )
                for row in (
                    expensive_input_df
                    .to_dict("records")
                )
            )
            scored_df = pd.DataFrame(
                scored_rows
            )
            if scored_df.empty:
                print(
                    "No molecule completed CE prediction; "
                    "sampling a fresh batch.",
                    flush=True,
                )
                current_smiles = sample_unique(
                    model,
                    BATCH_SIZE,
                    exclude=previously_seen,
                )
                previously_seen.update(
                    current_smiles
                )
                continue
            if "Predicted_CE" not in scored_df.columns:
                print(
                    "No Predicted_CE column was produced; "
                    "sampling a fresh batch.",
                    flush=True,
                )
                current_smiles = sample_unique(
                    model,
                    BATCH_SIZE,
                    exclude=previously_seen,
                )
                previously_seen.update(
                    current_smiles
                )
                continue
            scored_df["Predicted_CE"] = pd.to_numeric(
                scored_df["Predicted_CE"],
                errors="coerce",
            )
            scored_df["SoftScore"] = pd.to_numeric(
                scored_df["SoftScore"],
                errors="coerce",
            )
            scored_df = scored_df.dropna(
                subset=[
                    "SMILES",
                    "Predicted_CE",
                    "SoftScore",
                ]
            ).copy()
            if scored_df.empty:
                print(
                    "Every CE prediction failed; "
                    "sampling a fresh batch.",
                    flush=True,
                )
                current_smiles = sample_unique(
                    model,
                    BATCH_SIZE,
                    exclude=previously_seen,
                )
                previously_seen.update(
                    current_smiles
                )
                continue
            # Store every valid candidate from this iteration.
            all_scored_molecules.extend(
                scored_df.to_dict("records")
            )
            print(f"Stored {len(scored_df)} valid molecules "
                f"from iteration {iteration}.",
                flush=True,)
            print(f"Total stored candidate records: "
                f"{len(all_scored_molecules)}",
                flush=True, )
            preferred_df = scored_df[
                scored_df["SoftScore"]
                >= MIN_SOFT_SCORE
            ].copy()
            if preferred_df.empty:
                print(
                    "No candidate reached the preferred soft score "
                    "this iteration. Using the highest-soft-score "
                    "hard-pass candidates instead; the loop will continue.",
                    flush=True,
                )
                preferred_df = scored_df.nlargest(
                    min(
                        TOP_K,
                        len(scored_df),
                    ),
                    "SoftScore",
                ).copy()
            preferred_df = (
                preferred_df
                .sort_values(
                    [
                        "Predicted_CE",
                        "SoftScore",
                    ],
                    ascending=[
                        False,
                        False,
                    ],
                )
                .reset_index(drop=True)
            )
            top_row = (
                preferred_df
                .iloc[0]
                .to_dict()
            )
            top_ce = float(
                top_row["Predicted_CE"]
            )
            print(
                "Stage 3/3: selecting parents for "
                "the next MoLeR generation...",
                flush=True,
            )
            print(
                f"Top molecule this iteration: "
                f"{top_row['SMILES']}\n"
                f"  Predicted CE: {top_ce:.2f}%\n"
                f"  Soft chemistry score: "
                f"{float(top_row['SoftScore']):.1f}/100\n"
                f"  Molecular weight: "
                f"{finite_float(top_row.get('MolWt'), float('nan')):.2f}\n"
                f"  LogP: "
                f"{finite_float(top_row.get('LogP'), float('nan')):.2f}\n"
                f"  TPSA: "
                f"{finite_float(top_row.get('TPSA'), float('nan')):.2f}\n"
                f"  ESOL logS: "
                f"{finite_float(top_row.get('ESOL_logS'), float('nan')):.2f}\n"
                f"  SA score: "
                f"{finite_float(top_row.get('SA_Score'), float('nan')):.2f}\n"
                f"  xTB success: "
                f"{bool(top_row.get('XTBSuccess', False))}",
                flush=True,
            )
            target_count = int(
                (
                    preferred_df["Predicted_CE"]
                    >= TARGET_CE_FOR_REPORTING
                ).sum()
            )
            print(
                f"Candidates at or above "
                f"{TARGET_CE_FOR_REPORTING:.1f}% predicted CE: "
                f"{target_count}/{len(preferred_df)}",
                flush=True,
            )
            parent_df = preferred_df.head(
                min(
                    TOP_K,
                    len(preferred_df),
                )
            )
            top_k_smiles = (
                parent_df["SMILES"]
                .dropna()
                .astype(str)
                .tolist()
            )
            if not top_k_smiles:
                print(
                    "No parent SMILES were available. "
                    "Sampling a new independent batch.",
                    flush=True,
                )
                current_smiles = sample_unique(
                    model,
                    BATCH_SIZE,
                    exclude=previously_seen,
                )
                previously_seen.update(
                    current_smiles
                )
                continue
            print(
                f"Selected {len(top_k_smiles)} "
                "parent molecules.",
                flush=True,
            )
            try:
                latents = model.encode(
                    top_k_smiles
                )
                if isinstance(latents, tuple):
                    latents = latents[0]
                latents = np.asarray(
                    latents,
                    dtype=np.float32,
                )
            except Exception as exc:
                print(
                    f"MoLeR encoding failed: {exc}",
                    flush=True,
                )
                current_smiles = sample_unique(
                    model,
                    BATCH_SIZE,
                    exclude=previously_seen,
                )
                previously_seen.update(
                    current_smiles
                )
                continue
            if (
                latents.ndim < 2
                or len(latents) == 0
            ):
                print(
                    "MoLeR returned invalid latent representations. "
                    "Sampling a fresh batch.",
                    flush=True,
                )
                current_smiles = sample_unique(
                    model,
                    BATCH_SIZE,
                    exclude=previously_seen,
                )
                previously_seen.update(
                    current_smiles
                )
                continue
            repeats = int(
                np.ceil(
                    BATCH_SIZE
                    / len(top_k_smiles)
                )
            )
            batch_latents = np.repeat(
                latents,
                repeats,
                axis=0,
            )[:BATCH_SIZE]
            if (
                previous_best_ce is not None
                and top_ce <= previous_best_ce
            ):
                current_noise = min(
                    current_noise + 0.20,
                    1.50,
                )
                print(
                    "No CE improvement detected; "
                    "next-iteration noise increased "
                    f"to {current_noise:.2f}.",
                    flush=True,
                )
            else:
                current_noise = BASE_NOISE_SCALE
            previous_best_ce = top_ce
            noise = RNG.normal(
                loc=0.0,
                scale=current_noise,
                size=batch_latents.shape,
            ).astype(np.float32)
            try:
                decoded = model.decode(
                    (
                        batch_latents
                        + noise
                    ).astype(np.float32)
                )
            except Exception as exc:
                print(
                    f"MoLeR decoding failed: {exc}",
                    flush=True,
                )
                decoded = []
            next_generation: Dict[str, None] = {}
            parent_set = set(top_k_smiles)
            for candidate in decoded:
                canonical = canonicalize_smiles(
                    candidate
                )
                if (
                    canonical
                    and canonical not in parent_set
                    and canonical not in previously_seen
                ):
                    next_generation[canonical] = None
            # Fill any missing positions using independent MoLeR sampling.
            refill_attempts = 0
            maximum_refill_attempts = 20
            while (
                len(next_generation) < BATCH_SIZE
                and refill_attempts
                < maximum_refill_attempts
            ):
                refill_attempts += 1
                fresh = sample_unique(
                    model,
                    max(
                        5,
                        BATCH_SIZE
                        - len(next_generation),
                    ),
                    exclude=(
                        previously_seen
                        | parent_set
                        | set(next_generation)
                    ),
                )
                if not fresh:
                    break
                for candidate in fresh:
                    next_generation[candidate] = None
            current_smiles = list(
                next_generation
            )[:BATCH_SIZE]
            previously_seen.update(
                current_smiles
            )
            print(
                f"Iteration {iteration}/{ITERATIONS} complete. "
                f"Prepared {len(current_smiles)} molecules "
                "for the next iteration.",
                flush=True,
            )
            if (
                iteration < ITERATIONS
                and not current_smiles
            ):
                print(
                    "No new unique molecules were produced. "
                    "Sampling an independent batch.",
                    flush=True,
                )
                current_smiles = sample_unique(
                    model,
                    BATCH_SIZE,
                    exclude=previously_seen,
                )
                previously_seen.update(
                    current_smiles
                )
    print(
        "\n" + "=" * 78,
        flush=True,
    )
    print(
        "RECURSIVE MOLER SEARCH COMPLETE",
        flush=True,
    )
    if not all_scored_molecules:
        print(
            "No molecule received a valid CE prediction.",
            flush=True,
        )
        return None
    # Select the overall top unique molecules.
    top_internal_df, top_excel_df = (
        prepare_top_molecules_table(
            all_scored_molecules,
            FINAL_TOP_N,
        )
    )
    if top_excel_df.empty:
        print(
            "No valid molecules were available "
            "for the final top-molecule table.",
            flush=True,
        )
        return None
    # Save the final Excel workbook.
    save_top_molecules_excel(
        top_excel_df,
        OUTPUT_EXCEL_PATH,
    )
    # Print every final molecule and its requested properties.
    print_top_molecules(
        top_excel_df
    )
    print(
        "\n" + "=" * 78,
        flush=True,
    )
    print(
        f"Top {len(top_excel_df)} unique molecules "
        "saved successfully.",
        flush=True,
    )
    print(
        f"Excel file: "
        f"{Path(OUTPUT_EXCEL_PATH).resolve()}",
        flush=True,
    )
    # Returning the table is useful when running the script in Jupyter.
    return top_excel_df
if __name__ == "__main__":
    top_20_results = main()
